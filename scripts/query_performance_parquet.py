import os
import pandas as pd
import duckdb
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from .add_column import add_grouping_late, add_aging_carrer, add_status_pod_2, fix_regional_cols, TRANSFORM_GROUPS, TRANSFORM_FUNCS
from .parse_dates import normalize_all_dates
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ------------------- HELPERS: DuckDB-based IO -------------------
def get_duckdb_connection():
    return duckdb.connect(database=":memory:", read_only=False)


def _load_filtered_parquet_with_duckdb(
    con,
    files,
    group_name=None,
    date_col="TGL_ENTRY",
    start_dt=None,
    end_dt=None,
    id_account=None,
    selected_statuses=None,
    categories=None,
    debug=False,
):
    if not files:
        return pd.DataFrame()

    try:
        where_clauses = []
        params = []

        # BIG_GROUPING_CUST
        if group_name is not None:
            if isinstance(group_name, (list, tuple, set)):
                placeholders = ", ".join(["?"] * len(group_name))
                where_clauses.append(f"upper(BIG_GROUPING_CUST) IN ({placeholders})")
                params.extend([str(g).upper() for g in group_name])
            else:
                where_clauses.append("upper(BIG_GROUPING_CUST) = ?")
                params.append(str(group_name).upper())

        # CATEGORY column filter (case-insensitive)
        # accept either a single value or iterable of values via `categories`
        if categories is not None:
            if isinstance(categories, (list, tuple, set)):
                placeholders = ", ".join(["?"] * len(categories))
                where_clauses.append(f"upper(CATEGORY) IN ({placeholders})")
                params.extend([str(c).upper() for c in categories])
            else:
                where_clauses.append("upper(CATEGORY) = ?")
                params.append(str(categories).upper())

        # date range
        if start_dt is not None:
            where_clauses.append(f'"{date_col}" >= ?')
            params.append(pd.Timestamp(start_dt).to_pydatetime())
        if end_dt is not None:
            where_clauses.append(f'"{date_col}" < ?')
            params.append(pd.Timestamp(end_dt).to_pydatetime())

        # ID_ACCOUNT
        if id_account:
            placeholders = ", ".join(["?"] * len(id_account))
            where_clauses.append(f'"ID_ACCOUNT" IN ({placeholders})')
            # keep leading single-quote in ID_ACCOUNT values (some data stores ID_ACCOUNT like "'80656700")
            params.extend([str(x) if str(x).startswith("'") else "'" + str(x) for x in id_account])

        # STATUS_POD (case-insensitive match)
        if selected_statuses:
            placeholders = ", ".join(["?"] * len(selected_statuses))
            where_clauses.append(f"upper(STATUS_POD) IN ({placeholders})")
            params.extend([str(x).upper() for x in selected_statuses])

        where_sql = ""
        if where_clauses:
            where_sql = "WHERE " + " AND ".join(where_clauses)

        sql = f"""
            SELECT *
            FROM read_parquet($1)
            {where_sql}
        """

        # quick total rows count (best-effort)
        try:
            total_rows = con.execute("SELECT COUNT(*) AS c FROM read_parquet($1)", [files]).fetchone()[0]
        except Exception:
            total_rows = None

        if debug:
            print("[DEBUG SQL]", sql)
            print("[DEBUG PARAMS]", params)
            print(f"[DEBUG] parquet total rows: {total_rows}")

            # show a small sample of rows and some distinct values to help debugging
            try:
                sample = con.execute("SELECT * FROM read_parquet($1) LIMIT 5", [files]).df()
                if not sample.empty:
                    print("[DEBUG SAMPLE ROWS]")
                    print(sample.head().to_dict(orient='records'))
            except Exception as e:
                print(f"[DEBUG] sample read failed: {e}")

            try:
                vals = con.execute("SELECT DISTINCT upper(BIG_GROUPING_CUST) AS v FROM read_parquet($1) LIMIT 50", [files]).df()
                if not vals.empty and 'v' in vals.columns:
                    print("[DEBUG] BIG_GROUPING_CUST values:", vals['v'].tolist())
            except Exception:
                pass

            try:
                ids = con.execute("SELECT DISTINCT ID_ACCOUNT AS a FROM read_parquet($1) LIMIT 50", [files]).df()
                if not ids.empty and 'a' in ids.columns:
                    print("[DEBUG] sample ID_ACCOUNT values:", ids['a'].tolist()[:10])
            except Exception:
                pass

        df = con.execute(sql, [files] + params).df()

        # If nothing matched and we filtered by BIG_GROUPING_CUST, retry without that filter
        if (df is None or df.empty) and group_name is not None:
            try:
                if debug:
                    print(f"[DEBUG] no rows after applying BIG_GROUPING_CUST for '{group_name}', retrying without that filter...")

                # remove BIG_GROUPING_CUST clauses from where_clauses
                new_where = [c for c in where_clauses if 'BIG_GROUPING_CUST' not in c.upper()]
                # need to rebuild params accordingly: remove first len(group_name) params if group_name was list, else first param
                new_params = list(params)
                if isinstance(group_name, (list, tuple, set)):
                    remove_n = len(group_name)
                else:
                    remove_n = 1
                # remove the earliest matching params (they were appended first)
                new_params = new_params[remove_n:]

                new_where_sql = ""
                if new_where:
                    new_where_sql = "WHERE " + " AND ".join(new_where)

                new_sql = f"""
                    SELECT *
                    FROM read_parquet($1)
                    {new_where_sql}
                """
                if debug:
                    print("[DEBUG SQL - fallback]", new_sql)
                    print("[DEBUG PARAMS - fallback]", new_params)

                df = con.execute(new_sql, [files] + new_params).df()
            except Exception as e:
                if debug:
                    print(f"[DEBUG] fallback query failed: {e}")

        if df is None or df.empty:
            if debug:
                print(f"[DEBUG] filtered result empty for file {files}")
            print(f"[DUCKDB] file={files} total_rows={total_rows} filtered_rows=0")
            return pd.DataFrame()

        df.columns = [c.strip().upper() for c in df.columns]

        # normalize CATEGORY values to uppercase so downstream logic sees consistent casing
        if 'CATEGORY' in df.columns:
            try:
                df['CATEGORY'] = df['CATEGORY'].astype(str).str.upper()
            except Exception:
                # best-effort: if conversion fails, leave values as-is
                pass

        try:
            filtered_count = len(df)
        except Exception:
            filtered_count = None

        if debug:
            print(f"[DEBUG] filtered rows: {filtered_count}")

        print(f"[DUCKDB] file={files} total_rows={total_rows} filtered_rows={filtered_count}")
        return df

    except Exception as e:
        try:
            print(f"❌ DuckDB filtered read failed: {e}")
            if debug:
                print("[DEBUG SQL when error]", sql)
                print("[DEBUG PARAMS when error]", params)
        except Exception:
            pass
        return pd.DataFrame()


def _collect_parquet_files(folders, base_dir=None, category=None, debug=False):
    files = []

    for folder in folders:
        try:
            if not os.path.exists(folder):
                if debug:
                    print(f"[DEBUG] folder not found: {folder}")
                continue

            folder_count = 0
            # files directly in folder
            for entry in os.listdir(folder):
                path = os.path.join(folder, entry)
                if os.path.isfile(path) and entry.lower().endswith(".parquet"):
                    files.append(path)
                    folder_count += 1
                elif os.path.isdir(path):
                    # one-level deep: check files inside this subfolder
                    try:
                        for sub in os.listdir(path):
                            subp = os.path.join(path, sub)
                            if os.path.isfile(subp) and sub.lower().endswith('.parquet'):
                                files.append(subp)
                                folder_count += 1
                        continue
                    except Exception:
                        pass
            if debug:
                print(f"[DEBUG] scanned folder: {folder} -> found {folder_count} parquet file(s)")
        except Exception:
            continue

    # fallback: limited recursive search under base_dir
    if not files and base_dir and category:
        if debug:
            print(f"[DEBUG] explicit search found nothing, attempting limited recursive scan under {base_dir} for category={category}")
        cat_low = category.lower()
        for root, _, filenames in os.walk(base_dir):
            root_low = root.replace("\\", "/").lower()
            if ("category=" in root_low and f"category={cat_low}" in root_low) or \
               (f"/category/{cat_low}" in root_low) or root_low.endswith(f"/{cat_low}"):
                for fname in filenames:
                    if fname.lower().endswith(".parquet"):
                        files.append(os.path.join(root, fname))

        if not files:
            if debug:
                print(f"[DEBUG] limited recursive scan found nothing; doing last-resort search for parent folder named '{category}'")
            for root, _, filenames in os.walk(base_dir):
                if os.path.basename(root).lower() == cat_low:
                    for fname in filenames:
                        if fname.lower().endswith(".parquet"):
                            files.append(os.path.join(root, fname))

    unique_files = sorted(set(files))
    if debug:
        print(f"[DEBUG] total parquet files collected: {len(unique_files)}")
    return unique_files


class ProjectProcessorParquet:
    progress_status = {}

    def __init__(self, project_lists, start_yy_mm, end_yy_mm, base_path, archive_dir,
                progress_dict=None, full=False, status=None, report=False, criteria_lists=None,
                start_date=None, end_date=None):
        # reuse most of the CSV-based interface
        self.project_lists = project_lists
        self.start_yy_mm = start_yy_mm
        self.end_yy_mm = end_yy_mm
        self.base_path = base_path
        self.archive_dir = archive_dir
        self.progress_dict = progress_dict
        self.full = full
        self.status = status
        self.report = report
        self.criteria_lists = criteria_lists or []
        self.start_date = start_date
        self.end_date = end_date

        self.bulan_id = {
            1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL",
            5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
            9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER"
        }

        # Minimal set used when not full/report
        self.columns_to_extract = [
            'AWB', 'ID_ACCOUNT',"SHIPPER_NAME", 'TGL_ENTRY', 'CONSIGNEE_NAME', 'NOREF', 'ORIGIN', 'DEST', 'SERVICE', 'QTY', 
            'WEIGHT', 'AMOUNT', 'CODING', 'TGL_RECEIVED', 'ETD', 'COD_FLAG', 'BILNOTE_FLAG', 'BILNOTE_AMOUNT', 
            'GROUPING_SHIPPER', 'PERIODE', 'PERIODE_WEEK', '3 LC DEST', 'NAMA KAB/KOTA 2', 'REGIONAL', 'ZONA', 
            'PAYMENT_METHODE', 'STATUS_POD_2', '1ST_ATTEMPT_DATE', 'AGING_1ST', 'CAREER_1ST', 'AGING_POD', 
            'CARRER_POD', 'CODING_UNDEL', 'REASON RETURN', 'GROUPING_LATE'
        ]

        self.renamed_columns = {
            'PERIODE_WEEK': 'PERIODE WEEK',
            'NAMA KAB/KOTA 2': 'DEST 3',
            '3 LC DEST': 'DEST2',
            'PAYMENT_METHODE': 'PAYMENT METHODE',
            '1ST_ATTEMPT_DATE': '1ST ATTEMPT',
            'AGING_1ST': 'AGING 1st ATTEMPT',
            'CAREER_1ST': 'CAREER 1st ATTEMPT',
            'CARRER_POD': 'CARRER POD',
            'AGING_POD': 'AGING POD',
            'CODING_UNDEL': 'CODING RETURN',
            'GROUPING_LATE': 'REASON LATE'
        }

        self.full_columns = [
            "AWB","ID_ACCOUNT","SHIPPER_NAME","TGL_ENTRY","CONSIGNEE_NAME","ADDR1","ADDR2","ADDR3","CONTACT","NOTELP",
            "NOREF","ORIGIN","DEST","SERVICE","QTY","WEIGHT","GOODS_DESCR","INSURANCE_ID","GOODS_VALUE",
            "INSURANCE_VALUE(+)","AMOUNT","INTRUCTION","NOTICE","HOLD_REASON","RECEIVING","RECEIVING_DATE",
            "OUTBOUND_MANIFEST","OUTBOUND_MANIFEST_DATE","INBOUND_MANIFEST","USER_IM","INBOUND_MANIFEST_DATE",
            "MANIFEST_TRANSIT_AGEN","DATE_TRANSIT","HVO_NO","HVO_DATE","HVO_HUB","HVO_HUB_NAME","HVO_HUB_DESTINATION",
            "HVO_HUB_DESTINATION_NAME","HVI_NO","HVI_DATE","RUNSHEET_NO","DATE_RUNSHEET","RUNSHEET_COURIER_ID",
            "RUNSHEET_COURIER_NAME","CODING","STATUS_POD","TGL_RECEIVED","STATUS_LATITUDE","STATUS_LONGITUDE","AGING",
            "ETD","SLA","CARRER","RECEIVED/REASON","TGL_UPDATE_STATUS_POD","WUS_OUTGOING_CODE","WUS_REMARKS","WUS_DATE",
            "INVOICED","AWB_CANCEL","COD_FLAG","BILNOTE_FLAG","BILNOTE_AMOUNT","REFNO_UOB","SCO_NO","WO/DO/PO",
            "NO_INVOICE","PAYMENT_TYPE","DATE_1ST_ATTEMPT","RESULT_1ST_ATTEMPT","LATLONG_1ST_ATTEMPT","DATE_2ND_ATTEMPT",
            "RESULT_2ND_ATTEMPT","LATLONG_2ND_ATTEMPT","DATE_LAST_ATTEMPT","RESULT_LAST_ATTEMPT","LATLONG_LAST_ATTEMPT",
            "PRA_RUNSHEET_NO","PRA_RUNSHEET_NAME","PRA_RUNSHEET_DATE","CS3_DATE","CONNOTE_RETURN_RT",
            "DATE_CONNOTE_RETURN_RT","CONNOTE_RETURN_RF","DATE_CONNOTE_RETURN_RF","USER_CONNOTE","USER_ZONE_CONNOTE",
            "CONFIRM_SHIPMENT_UNDEL","TRANSIT_MANIFEST","TRANSIT_MANIFEST_DATE","TRANSIT_MANIFEST_USER","IREG_MANIFEST",
            "IREG_CODE","IREG_DATE","URL_TTD","URL_FOTO","USER_OM","USER_RECEIVING","AGING_ONGOING","CLAIM_NO",
            "CLAIM_DOC_NO","CLAIM_DATE","NO_CNOTE_FW","ORIGIN_FW","DEST_FW","CODING_STATUS_FW","DESC_STATUS_FW","HBG_NO",
            "HBG_DATE","1ST_HVO_NO","1ST_HVO_DATE","1ST_HVO_USER","LAST_HVO_NO","LAST_HVO_DATE","LAST_HVO_USER",
            "MANIFEST_TRANSIT_SUBAGEN_NO","MANIFEST_TRANSIT_SUBAGEN_DATE","MANIFEST_INBOUND_SUBAGEN_NO",
            "MANIFEST_INBOUND_SUBAGEN_DATE","BAG_NO","LATEST_SM_NO","LATEST_SM_DATE","1ST_PREVIOUS_SM_NO",
            "1ST_PREVIOUS_SM_DATE","2ND_PREVIOUS_SM_NO","2ND_PREVIOUS_SM_DATE","1ST_TRANSIT_MANIFEST_NO",
            "1ST_TRANSIT_MANIFEST_DATE","2ND_TRANSIT_MANIFEST_NO","2ND_TRANSIT_MANIFEST_DATE","3RD_TRANSIT_MANIFEST_NO",
            "3RD_TRANSIT_MANIFEST_DATE","LAST_TRANSIT_MANIFEST_NO","LAST_TRANSIT_MANIFEST_DATE","MTI_USER","MTS_USER",
            "HO_COURIER_NO","HO_COURIER_DATE","WAREHOUSE_DATE","OFFICE_DATE","IRREG_REMAKS","BPIK","ZONE_USER_ENTRI",
            "CORRECT_DESTINATION","CORRECT_SERVICE","CORRECT_AMOUNT","HACB_NO","HACB_DATE","HACB_USER","HBAG_NO",
            "HBAG_DATE","HBAG_USER","PICKUP_DATE","PICKUP_STATUS","PICKUP_COURIER_ID","1ST_RUNSHEET_DATE",
            "1ST_RUNSHEET_COURIERID","URL_CHAT","SINGLE_LEG","LAST_DATE_DO","NO_RCW","DATE_RCW","USER_RCW","DATE_LPR",
            "NO_LPR","NO_RDO","DATE_RDO","NO_DO","PROJECT_KR","HO_OFFICE_NO","HO_OFFICE_DATE","LATEST_SM_ORIGIN",
            "LATEST_SM_DEST","1ST_PREVIOUS_SM_ORIGIN","1ST_PREVIOUS_SM_DEST","2ND_PREVIOUS_SM_ORIGIN",
            "2ND_PREVIOUS_SM_DEST","TGL_TARIK_REPORT","RESPONCIBILITY","STATUS_VERSI_CCC","STATUS_POD_UPDATE",
            "1ST_ATTEMPT_DATE","AGING_1ST","CAREER_1ST","AGING_POD","CARRER_POD","CODING_UNDEL","REASON RETURN",
            "3 LC DEST","NAMA KAB/KOTA 2","REGIONAL","ZONA","GROUPING_SHIPPER","CATEGORY","REFERENCE CUST CCC",
            "PAYMENT_METHODE","CUST_INDUSTRY","BIG_GROUPING_CUST","PIC_NAME_NEW RELATION","PIC SUPPORT DATA","UNIT",
            "DEPT","DATE","PERIODE","PERIODE_WEEK","ORIGIN RT","DEST RT","3LC Last Status","Regional Last Status",
            "Zona Last Status","OTS by CT","PERIODE_OTS","CLOSING_OTS","CODING_RT","RECEIVED/REASON_RT","Return Date"
        ]

    def parse_yy_mm(self, inp):
        tahun = int("20" + inp[:2])
        bulan = int(inp[2:])
        return tahun, bulan

    def month_iter(self, start, end):
        cur = start
        while cur <= end:
            yield cur
            if cur.month == 12:
                cur = datetime(cur.year + 1, 1, 1)
            else:
                cur = datetime(cur.year, cur.month + 1, 1)

    def get_criteria_for_project(self, proj_name):
        if not self.criteria_lists:
            return None
        for crit in self.criteria_lists:
            if crit["group_name"].strip().lower() == proj_name.strip().lower():
                return crit
        return None

    def run(self, request_id=None, filter_categories=None, debug=False):
        start_year, start_month = self.parse_yy_mm(self.start_yy_mm)
        end_year, end_month = self.parse_yy_mm(self.end_yy_mm)

        if self.start_date and self.end_date:
            start_date = self.start_date
            start_date_iter = datetime(self.start_date.year, self.start_date.month, 1)
            end_date_iter = datetime(self.end_date.year, self.end_date.month, 1)
            end_date = self.end_date + timedelta(days=1)
        else:
            start_date = datetime(start_year, start_month, 1)
            start_date_iter = start_date
            # end_date untuk iterasi folder = akhir bulan yang diminta
            end_date_iter = datetime(end_year, end_month, 1)
            # end_date untuk filter data = awal bulan berikutnya (eksklusif)
            end_date = datetime(end_year, end_month, 1) + relativedelta(months=1)
        is_multi_project = len(self.project_lists) > 1

        if self.full:
            if filter_categories:
                categories = [c.lower().strip() for c in filter_categories]
            else:
                categories = list({proj["category"].lower() for proj in self.project_lists})
        else:
            categories = list({proj["category"].lower() for proj in self.project_lists})

        project_id_map = {}
        for proj in self.project_lists:
            vals = []
            for x in proj.get("id_account", []):
                s = str(x).strip()
                if not s.startswith("'"):
                    s = "'" + s
                vals.append(s)
            project_id_map[proj["name"]] = vals

        results = {proj["name"]: [] for proj in self.project_lists}

        STAGE_WEIGHTS = {"reading": 0.4, "processing": 0.4, "saving": 0.2}

        if filter_categories:
            filter_categories = [c.lower().strip() for c in filter_categories]

        # collect all parquet files across months/categories first
        all_files = []
        for cur_date in self.month_iter(start_date_iter, end_date_iter):
            year, month = cur_date.year, cur_date.month
            month_name = self.bulan_id[month]
            for category in categories:
                if filter_categories and category.lower() not in filter_categories:
                    continue

                legacy = os.path.join(self.base_path, str(year), f"{month:02d}. {month_name} {year}", "CATEGORY", category.upper())
                partition = os.path.join(self.base_path, f"tahun={year}", f"bulan={month:02d}", f"category={category}")
                files = _collect_parquet_files([legacy, partition], base_dir=self.base_path, category=category, debug=debug)
                if files:
                    all_files.extend(files)

        all_files = sorted(set(all_files))
        total_files = len(all_files)
        if total_files == 0:
            print("⚠ Tidak ada file Parquet ditemukan.")
            if request_id and self.progress_dict is not None:
                self.progress_dict[request_id] = {"current": 100, "total": 100}
            return []

        # simple estimation for chunks: 1 chunk per file (DuckDB returns already filtered rows)
        chunksize = 50000
        total_chunks = total_files

        if request_id and self.progress_dict is not None:
            self.progress_dict[request_id] = {"current": 0, "total": 100}

        con = get_duckdb_connection()

        processed_files = 0
        processed_chunks = 0

        for file_path in all_files:
            if debug:
                print(f"   ⏳ Baca parquet: {file_path}")
            # --- If FULL mode with a category filter, load file-level CATEGORY-filtered data
            if self.full and filter_categories:
                try:
                    df_all = _load_filtered_parquet_with_duckdb(
                        con,
                        file_path,
                        group_name=None,
                        date_col="TGL_ENTRY",
                        start_dt=start_date,
                        end_dt=end_date,
                        id_account=None,
                        selected_statuses=None,
                        categories=filter_categories,
                        debug=debug,
                    )
                except Exception as e:
                    print(f"Error loading parquet (category mode) {file_path}: {e}")
                    df_all = pd.DataFrame()

                if df_all is None or df_all.empty:
                    if debug:
                        print(f"[DEBUG] no rows returned for category filter from {file_path}")
                else:
                    results.setdefault("ALL_CATEGORIES", []).append(df_all)
                    processed_chunks += 1

                # move to next file (we don't split per-project in this mode)
                continue

            for proj in self.project_lists:
                proj_name = proj["name"]
                id_accounts_clean = project_id_map[proj_name]

                try:
                    criteria = self.get_criteria_for_project(proj_name)
                    sel_statuses = None
                    if criteria and "selected_statuses" in criteria:
                        sel_statuses = criteria.get("selected_statuses")

                    df_proj = _load_filtered_parquet_with_duckdb(
                        con,
                        file_path,
                        group_name=proj_name,
                        date_col="TGL_ENTRY",
                        start_dt=start_date,
                        end_dt=end_date,
                        id_account=id_accounts_clean,
                        selected_statuses=sel_statuses,
                        debug=debug,
                    )
                except Exception as e:
                    print(f"Error loading parquet {file_path} for project {proj_name}: {e}")
                    continue

                if df_proj is None or df_proj.empty:
                    if debug:
                        print(f"[DEBUG] no rows returned for project '{proj_name}' from {file_path}")
                    continue

                if debug:
                    try:
                        print(f"[DEBUG] loaded {len(df_proj)} rows for project '{proj_name}' from {file_path}")
                        print("[DEBUG] columns:", list(df_proj.columns)[:20])
                    except Exception:
                        pass

                # split into chunks to mimic CSV chunk processing
                for start in range(0, len(df_proj), chunksize):
                    chunk = df_proj.iloc[start:start+chunksize].copy()

                    if proj.get("late_delivery", False):
                        chunk = normalize_all_dates(chunk, debug=False)
                        chunk = add_grouping_late(chunk)

                    if not self.report:
                        chunk = add_status_pod_2(chunk, debug=False)
                    if 'STATUS_POD_2' not in chunk.columns:
                        chunk['STATUS_POD_2'] = 'Unknown'

                    if self.status and self.status.lower() != "all":
                        if self.status.lower() == "close":
                            chunk = chunk[chunk["STATUS_POD_2"].isin(["Success", "Return Shipper"]) ]
                        elif self.status.lower() == "open":
                            chunk = chunk[~chunk["STATUS_POD_2"].isin(["Success", "Return Shipper"]) ]

                    if not self.full and not self.report:
                        cols = [col for col in self.columns_to_extract if col in chunk.columns]
                        chunk = chunk[cols]

                    results[proj_name].append(chunk)
                    processed_chunks += 1

            processed_files += 1
            if request_id and self.progress_dict is not None:
                reading_progress = (processed_files / total_files) * STAGE_WEIGHTS["reading"] * 100
                processing_progress = (processed_chunks / total_chunks) * STAGE_WEIGHTS["processing"] * 100
                current_progress = reading_progress + processing_progress
                self.progress_dict[request_id]["current"] = min(round(current_progress), 99)

        con.close()

        # reuse remaining save logic from CSV-based processor (concatenate, dedupe, save)
        MAX_ROWS_EXCEL = 1_000_000
        saved_files = []
        processed_saves = 0

        total_projects = len(results)
        total_saves = 0
        for proj_name, df_list in results.items():
            if not df_list:
                continue
            df_final = pd.concat(df_list, ignore_index=True)
            if "AWB" in df_final.columns:
                if "TGL_TARIK_REPORT" in df_final.columns:
                    df_final["TGL_TARIK_REPORT"] = pd.to_datetime(df_final["TGL_TARIK_REPORT"], errors="coerce")
                    df_final = df_final.sort_values(by=["AWB", "TGL_TARIK_REPORT"], ascending=[True, False])
                df_final = df_final.drop_duplicates(subset=["AWB"], keep="last")
            total_rows = len(df_final)
            total_saves += (total_rows // MAX_ROWS_EXCEL) + (1 if total_rows % MAX_ROWS_EXCEL else 0)
        if total_saves == 0:
            total_saves = total_projects

        for proj_name, df_list in results.items():
            if not df_list:
                print(f"⚠ Tidak ada data untuk project {proj_name}")
                continue

            df_final = pd.concat(df_list, ignore_index=True)

            if "AWB" in df_final.columns:
                if "TGL_TARIK_REPORT" in df_final.columns:
                    df_final["TGL_TARIK_REPORT"] = pd.to_datetime(df_final["TGL_TARIK_REPORT"], errors="coerce")
                    df_final = df_final.sort_values(by=["AWB", "TGL_TARIK_REPORT"], ascending=[True, False])
                df_final = df_final.drop_duplicates(subset=["AWB"], keep="last")

            # df_final = normalize_all_dates(df_final, debug=False)

            if not self.full:
                criteria = self.get_criteria_for_project(proj_name)
                if criteria and "selected_cols" in criteria:
                    selected_cols = criteria["selected_cols"]
                    needed_cols = set(selected_cols)
                    for tg_name, group in TRANSFORM_GROUPS.items():
                        if any(col in needed_cols for col in group["cols"]):
                            try:
                                df_final = group["func"](df_final, self.base_path)
                            except Exception as e:
                                print(f"⚠️ Gagal transform group {tg_name}: {e}")

                    for col in needed_cols:
                        if col in TRANSFORM_FUNCS:
                            try:
                                df_final = TRANSFORM_FUNCS[col](df_final, self.base_path)
                            except Exception as e:
                                print(f"⚠️ Gagal transform func {col}: {e}")

                    for col in selected_cols:
                        if col not in df_final.columns:
                            df_final[col] = ""

                    df_final = df_final[[c for c in selected_cols if c in df_final.columns]]
                else:
                    print(f"⚠️ Criteria tidak ditemukan untuk {proj_name}, skip transform.")

            if filter_categories:
                print(f"ℹ️ Mode kategori-only — lewati aging_carrer & rename kolom.")
            elif self.criteria_lists:
                print(f"ℹ️ Mode report aktif — lewati aging_carrer.")
            else:
                df_final = add_aging_carrer(df_final, debug=debug)
                df_final.rename(columns=self.renamed_columns, inplace=True)

            df_final = fix_regional_cols(df_final)

            subfolder = os.path.join(self.archive_dir, proj_name)
            os.makedirs(subfolder, exist_ok=True)

            total_rows = len(df_final)
            num_parts = (total_rows // MAX_ROWS_EXCEL) + (1 if total_rows % MAX_ROWS_EXCEL else 0)

            safe_proj_name = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in proj_name)

            for part in range(num_parts):
                start = part * MAX_ROWS_EXCEL
                end = min((part + 1) * MAX_ROWS_EXCEL, total_rows)
                df_part = df_final.iloc[start:end]

                if num_parts > 1:
                    suffix = f" part {part + 1}"
                else:
                    suffix = ""
                filename = f"{safe_proj_name}{suffix} {self.start_yy_mm}-{self.end_yy_mm}"

                if self.full:
                    # 🔹 Simpan sebagai CSV
                    save_path = os.path.join(subfolder, filename + ".csv")
                    try:
                        # Pastikan semua kolom di full_columns ada
                        for col in self.full_columns:
                            if col not in df_part.columns:
                                df_part[col] = pd.NA

                        # Urutkan kolom sesuai definisi full_columns
                        df_part = df_part[self.full_columns]

                        df_part.to_csv(save_path, index=False, encoding="utf-8-sig")
                        print(f"✅ [FULL MODE] Project {proj_name} part {part+1}/{num_parts} disimpan ke CSV: {save_path}")
                        saved_files.append(save_path)
                    except Exception as e:
                        print(f"Error saving CSV {save_path}: {e}")
                        continue
                else:
                    save_path = os.path.join(subfolder, filename + ".xlsx")
                    try:
                        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                            df_part.to_excel(writer, sheet_name=proj_name, index=False)

                        # Avoid loading large workbooks into memory for styling
                        if len(df_part) <= 200000:
                            wb = openpyxl.load_workbook(save_path)
                            ws = wb[proj_name]
                            nrows, ncols = ws.max_row, ws.max_column
                            last_col = get_column_letter(ncols)
                            table_range = f"A1:{last_col}{nrows}"

                            if not self.report:
                                table = Table(displayName="DATA_DETAIL", ref=table_range)
                                style = TableStyleInfo(
                                    name="TableStyleMedium9",
                                    showFirstColumn=False,
                                    showLastColumn=False,
                                    showRowStripes=True,
                                    showColumnStripes=False
                                )
                                table.tableStyleInfo = style
                                ws.add_table(table)
                            else:
                                thin_border = Border(
                                    left=Side(style="thin"),
                                    right=Side(style="thin"),
                                    top=Side(style="thin"),
                                    bottom=Side(style="thin")
                                )
                                header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

                                for row in ws.iter_rows():
                                    for cell in row:
                                        cell.border = thin_border
                                        if cell.row == 1:
                                            cell.fill = header_fill
                                        if isinstance(cell.value, (datetime, pd.Timestamp)):
                                            cell.number_format = "MM/DD/YYYY"

                                for col_cells in ws.columns:
                                    col_letter = get_column_letter(col_cells[0].column)
                                    max_length = max(len(str(cell.value) if cell.value else "") for cell in col_cells)
                                    ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

                            wb.save(save_path)
                        else:
                            print("ℹ️ Data besar, skip styling Excel untuk hemat memori.")
                        print(f"✅ Project {proj_name} part {part+1}/{num_parts} disimpan ke Excel: {save_path}")
                        saved_files.append(save_path)
                    except Exception as e:
                        print(f"❌ Gagal menyimpan file {save_path}: {e}")
                        continue

                processed_saves += 1
                if request_id and self.progress_dict is not None:
                    saving_progress = (processed_saves / total_saves) * STAGE_WEIGHTS["saving"] * 100
                    current_progress = (
                        (processed_files / total_files) * STAGE_WEIGHTS["reading"] * 100 +
                        (processed_chunks / total_chunks) * STAGE_WEIGHTS["processing"] * 100 +
                        saving_progress
                    )
                    cap = 99 if is_multi_project else 100
                    self.progress_dict[request_id]["current"] = min(round(current_progress), cap)

        if request_id and self.progress_dict is not None:
            if is_multi_project:
                # wait until archive / winrar step
                self.progress_dict[request_id]["current"] = 99
            else:
                # single project → done immediately after save
                self.progress_dict[request_id]["current"] = 100

        print("🎉 All data processed and saved successfully.")
        return saved_files
