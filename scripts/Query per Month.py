import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from lists import project_lists 

# Mapping bulan ke bahasa Indonesia (uppercase)
bulan_id = {
    1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL",
    5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER"
}

# --- Input offset bulan ---
try:
    offset = int(input("Masukkan jumlah bulan ke belakang (contoh: 3 untuk 3 bulan lalu): "))
except ValueError:
    print("⚠ Input tidak valid, default ke 1 bulan lalu.")
    offset = 1

# Hitung bulan target
today = datetime.today()
first_day_this_month = today.replace(day=1)
target_date = first_day_this_month - pd.DateOffset(months=offset)

year = target_date.year
month = target_date.month
month_name = bulan_id[month]
periode_code = f"{str(year)[2:]}{month:02d}"  # contoh: 2507

print(f"📅 Mengambil data untuk: {month_name} {year} (Periode Code: {periode_code})")

# Dynamic folder path
base_path = os.getenv("CSV_BASE_PATH", "")
folder_base = fr"{base_path}\{year}\{month}. {month_name} {year}\CATEGORY"

# Output & Archive
archive_dir = os.getenv("QUERY_ARCHIVE_DIR", r'D:\RYAN\4. Performance Report\Data Archive')

columns_to_extract = [
    'AWB', 'ID_ACCOUNT', 'TGL_ENTRY', 'CONSIGNEE_NAME', 'NOREF', 'ORIGIN', 'DEST', 'SERVICE', 'QTY', 
    'WEIGHT', 'AMOUNT', 'CODING', 'TGL_RECEIVED', 'ETD', 'COD_FLAG', 'BILNOTE_FLAG', 'BILNOTE_AMOUNT', 
    'GROUPING_SHIPPER', 'PERIODE', 'PERIODE_WEEK', '3 LC DEST', 'NAMA KAB/KOTA 2', 'REGIONAL', 'ZONA', 
    'PAYMENT_METHODE', 'STATUS_POD_UPDATE', '1ST_ATTEMPT_DATE', 'AGING_1ST', 'CAREER_1ST', 'AGING_POD', 
    'CARRER_POD', 'CODING_UNDEL', 'REASON RETURN'
]

def combine_csv(folder_path):
    df_combined = pd.DataFrame()
    if not os.path.exists(folder_path):
        print(f"⚠ Folder tidak ditemukan: {folder_path}")
        return df_combined
    
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.csv'):
            file_path = os.path.join(folder_path, file_name)
            df = pd.read_csv(file_path, low_memory=False)
            cols = [col for col in columns_to_extract if col in df.columns]
            df_filtered = df[cols]
            df_combined = pd.concat([df_combined, df_filtered], ignore_index=True)
    return df_combined

def split_and_save(df, projects, archive_dir, periode_code):
    renamed_columns = {
        'PERIODE_WEEK': 'PERIODE WEEK',
        'NAMA KAB/KOTA 2': 'DEST 3',
        '3 LC DEST': 'DEST2',
        'PAYMENT_METHODE': 'PAYMENT METHODE',
        'STATUS_POD_UPDATE': 'STATUS_POD_2',
        '1ST_ATTEMPT_DATE': '1ST ATTEMPT',
        'AGING_1ST': 'AGING 1st ATTEMPT',
        'CAREER_1ST': 'CAREER 1st ATTEMPT',
        'CARRER_POD': 'CARRER POD',
        'AGING_POD': 'AGING POD',
        'CODING_UNDEL': 'CODING RETURN'
    }

    for project in projects:
        name = project["name"]
        id_accounts = project.get("id_account", [])
        id_accounts_clean = [str(x).replace("'", "").strip() for x in id_accounts]
        df_split = df[df['ID_ACCOUNT'].astype(str).str.replace("'", "").isin(id_accounts_clean)]

        if not df_split.empty:
            df_filtered = df_split.loc[:, df_split.columns.intersection(columns_to_extract)].copy()
            df_filtered.rename(columns=renamed_columns, inplace=True)

            # Konversi date-like columns
            for col in df_filtered.columns:
                col_lower = col.lower()
                if ('tgl' in col_lower or 'date' in col_lower) and col != "STATUS_POD_2":
                    try:
                        df_filtered[col] = pd.to_datetime(df_filtered[col], errors='coerce').dt.date
                    except Exception as e:
                        print(f"[!] Gagal konversi kolom {col} ke date: {e}")

            subfolder = os.path.join(archive_dir, name)
            os.makedirs(subfolder, exist_ok=True)

            filename = f"{name} {periode_code}.xlsx"
            save_path = os.path.join(subfolder, filename)

            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                df_filtered.to_excel(writer, sheet_name=name, index=False)

            wb = openpyxl.load_workbook(save_path)
            ws = wb[name]
            nrows = ws.max_row
            ncols = ws.max_column
            last_col = openpyxl.utils.get_column_letter(ncols)
            table_range = f"A1:{last_col}{nrows}"

            table = Table(displayName="DATA_DETAIL", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            wb.save(save_path)

# --- Process per category ---
categories = ["non big", "banking", "aggregator", "sca"]

for category in categories:
    folder_path = os.path.join(folder_base, category.upper())
    print(f"▶ Proses category: {category} ({folder_path})")

    df_category = combine_csv(folder_path)
    if df_category.empty:
        print(f"⚠ Tidak ada data di {category}")
        continue

    # Simpan CSV gabungan (opsional)
    output_path = os.path.join(os.getenv("QUERY_OUTPUT_DIR", "output"), f"{category.title()} Data Combined.csv")
    df_category.to_csv(output_path, index=False)

    # Filter hanya project dengan category tsb
    project = [e for e in project_lists if e["category"].lower() == category]
    split_and_save(df_category, project, archive_dir, periode_code)

print("✅ All data processed and saved successfully.")