import os
import codecs
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from .add_column import add_grouping_late, add_aging_carrer, add_status_pod_2, fix_regional_cols, TRANSFORM_GROUPS, TRANSFORM_FUNCS
from .parse_dates import normalize_all_dates


CSV_ENCODINGS_TO_TRY = ("utf-8-sig", "utf-8", "cp1252", "latin1")


def _can_decode_file(file_path, encoding, block_size=1024 * 1024):
    decoder = codecs.getincrementaldecoder(encoding)()
    with open(file_path, "rb") as file:
        while True:
            block = file.read(block_size)
            if not block:
                decoder.decode(b"", final=True)
                return True
            decoder.decode(block)


def _detect_csv_encoding(file_path):
    for encoding in CSV_ENCODINGS_TO_TRY:
        try:
            if _can_decode_file(file_path, encoding):
                return encoding
        except UnicodeDecodeError:
            continue
    return "latin1"


def _read_csv(file_path, **kwargs):
    encoding = _detect_csv_encoding(file_path)
    try:
        return pd.read_csv(file_path, encoding=encoding, **kwargs)
    except UnicodeDecodeError:
        for fallback_encoding in CSV_ENCODINGS_TO_TRY:
            if fallback_encoding == encoding:
                continue
            try:
                return pd.read_csv(file_path, encoding=fallback_encoding, **kwargs)
            except UnicodeDecodeError:
                continue
    return pd.read_csv(file_path, encoding="latin1", **kwargs)


class ProjectProcessor:
    progress_status = {}

    def __init__(self, project_lists, start_yy_mm, end_yy_mm, base_path, archive_dir,
                progress_dict=None, full=False, status=None, report=False, criteria_lists=None,
                start_date=None, end_date=None):
        self.project_lists = project_lists
        self.start_yy_mm = start_yy_mm
        self.end_yy_mm = end_yy_mm
        self.base_path = base_path
        self.archive_dir = archive_dir
        self.progress_dict = progress_dict
        self.full = full
        self.status = status
        self.report = report
        self.criteria_lists = criteria_lists or []
        self.start_date = start_date
        self.end_date = end_date

        self.bulan_id = {
            1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL",
            5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
            9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER"
        }
        self.columns_to_extract = [
            'AWB', 'ID_ACCOUNT', 'TGL_ENTRY', 'CONSIGNEE_NAME', 'NOREF', 'ORIGIN', 'DEST', 'SERVICE', 'QTY', 
            'WEIGHT', 'AMOUNT', 'CODING', 'TGL_RECEIVED', 'ETD', 'COD_FLAG', 'BILNOTE_FLAG', 'BILNOTE_AMOUNT', 
            'GROUPING_SHIPPER', 'PERIODE', 'PERIODE_WEEK', '3 LC DEST', 'NAMA KAB/KOTA 2', 'REGIONAL', 'ZONA', 
            'PAYMENT_METHODE', 'STATUS_POD_2', '1ST_ATTEMPT_DATE', 'AGING_1ST', 'CAREER_1ST', 'AGING_POD', 
            'CARRER_POD', 'CODING_UNDEL', 'REASON RETURN', 'GROUPING_LATE'
        ]
        self.renamed_columns = {
            'PERIODE_WEEK': 'PERIODE WEEK',
            'NAMA KAB/KOTA 2': 'DEST 3',
            '3 LC DEST': 'DEST2',
            'PAYMENT_METHODE': 'PAYMENT METHODE',
            '1ST_ATTEMPT_DATE': '1ST ATTEMPT',
            'AGING_1ST': 'AGING 1st ATTEMPT',
            'CAREER_1ST': 'CAREER 1st ATTEMPT',
            'CARRER_POD': 'CARRER POD',
            'AGING_POD': 'AGING POD',
            'CODING_UNDEL': 'CODING RETURN',
            'GROUPING_LATE': 'REASON LATE'
        }
        self.full_columns = [
            "AWB","ID_ACCOUNT","SHIPPER_NAME","TGL_ENTRY","CONSIGNEE_NAME","ADDR1","ADDR2","ADDR3","CONTACT","NOTELP",
            "NOREF","ORIGIN","DEST","SERVICE","QTY","WEIGHT","GOODS_DESCR","INSURANCE_ID","GOODS_VALUE",
            "INSURANCE_VALUE(+)","AMOUNT","INTRUCTION","NOTICE","HOLD_REASON","RECEIVING","RECEIVING_DATE",
            "OUTBOUND_MANIFEST","OUTBOUND_MANIFEST_DATE","INBOUND_MANIFEST","USER_IM","INBOUND_MANIFEST_DATE",
            "MANIFEST_TRANSIT_AGEN","DATE_TRANSIT","HVO_NO","HVO_DATE","HVO_HUB","HVO_HUB_NAME","HVO_HUB_DESTINATION",
            "HVO_HUB_DESTINATION_NAME","HVI_NO","HVI_DATE","RUNSHEET_NO","DATE_RUNSHEET","RUNSHEET_COURIER_ID",
            "RUNSHEET_COURIER_NAME","CODING","STATUS_POD","TGL_RECEIVED","STATUS_LATITUDE","STATUS_LONGITUDE","AGING",
            "ETD","SLA","CARRER","RECEIVED/REASON","TGL_UPDATE_STATUS_POD","WUS_OUTGOING_CODE","WUS_REMARKS","WUS_DATE",
            "INVOICED","AWB_CANCEL","COD_FLAG","BILNOTE_FLAG","BILNOTE_AMOUNT","REFNO_UOB","SCO_NO","WO/DO/PO",
            "NO_INVOICE","PAYMENT_TYPE","DATE_1ST_ATTEMPT","RESULT_1ST_ATTEMPT","LATLONG_1ST_ATTEMPT","DATE_2ND_ATTEMPT",
            "RESULT_2ND_ATTEMPT","LATLONG_2ND_ATTEMPT","DATE_LAST_ATTEMPT","RESULT_LAST_ATTEMPT","LATLONG_LAST_ATTEMPT",
            "PRA_RUNSHEET_NO","PRA_RUNSHEET_NAME","PRA_RUNSHEET_DATE","CS3_DATE","CONNOTE_RETURN_RT",
            "DATE_CONNOTE_RETURN_RT","CONNOTE_RETURN_RF","DATE_CONNOTE_RETURN_RF","USER_CONNOTE","USER_ZONE_CONNOTE",
            "CONFIRM_SHIPMENT_UNDEL","TRANSIT_MANIFEST","TRANSIT_MANIFEST_DATE","TRANSIT_MANIFEST_USER","IREG_MANIFEST",
            "IREG_CODE","IREG_DATE","URL_TTD","URL_FOTO","USER_OM","USER_RECEIVING","AGING_ONGOING","CLAIM_NO",
            "CLAIM_DOC_NO","CLAIM_DATE","NO_CNOTE_FW","ORIGIN_FW","DEST_FW","CODING_STATUS_FW","DESC_STATUS_FW","HBG_NO",
            "HBG_DATE","1ST_HVO_NO","1ST_HVO_DATE","1ST_HVO_USER","LAST_HVO_NO","LAST_HVO_DATE","LAST_HVO_USER",
            "MANIFEST_TRANSIT_SUBAGEN_NO","MANIFEST_TRANSIT_SUBAGEN_DATE","MANIFEST_INBOUND_SUBAGEN_NO",
            "MANIFEST_INBOUND_SUBAGEN_DATE","BAG_NO","LATEST_SM_NO","LATEST_SM_DATE","1ST_PREVIOUS_SM_NO",
            "1ST_PREVIOUS_SM_DATE","2ND_PREVIOUS_SM_NO","2ND_PREVIOUS_SM_DATE","1ST_TRANSIT_MANIFEST_NO",
            "1ST_TRANSIT_MANIFEST_DATE","2ND_TRANSIT_MANIFEST_NO","2ND_TRANSIT_MANIFEST_DATE","3RD_TRANSIT_MANIFEST_NO",
            "3RD_TRANSIT_MANIFEST_DATE","LAST_TRANSIT_MANIFEST_NO","LAST_TRANSIT_MANIFEST_DATE","MTI_USER","MTS_USER",
            "HO_COURIER_NO","HO_COURIER_DATE","WAREHOUSE_DATE","OFFICE_DATE","IRREG_REMAKS","BPIK","ZONE_USER_ENTRI",
            "CORRECT_DESTINATION","CORRECT_SERVICE","CORRECT_AMOUNT","HACB_NO","HACB_DATE","HACB_USER","HBAG_NO",
            "HBAG_DATE","HBAG_USER","PICKUP_DATE","PICKUP_STATUS","PICKUP_COURIER_ID","1ST_RUNSHEET_DATE",
            "1ST_RUNSHEET_COURIERID","URL_CHAT","SINGLE_LEG","LAST_DATE_DO","NO_RCW","DATE_RCW","USER_RCW","DATE_LPR",
            "NO_LPR","NO_RDO","DATE_RDO","NO_DO","PROJECT_KR","HO_OFFICE_NO","HO_OFFICE_DATE","LATEST_SM_ORIGIN",
            "LATEST_SM_DEST","1ST_PREVIOUS_SM_ORIGIN","1ST_PREVIOUS_SM_DEST","2ND_PREVIOUS_SM_ORIGIN",
            "2ND_PREVIOUS_SM_DEST","TGL_TARIK_REPORT","RESPONCIBILITY","STATUS_VERSI_CCC","STATUS_POD_UPDATE",
            "1ST_ATTEMPT_DATE","AGING_1ST","CAREER_1ST","AGING_POD","CARRER_POD","CODING_UNDEL","REASON RETURN",
            "3 LC DEST","NAMA KAB/KOTA 2","REGIONAL","ZONA","GROUPING_SHIPPER","CATEGORY","REFERENCE CUST CCC",
            "PAYMENT_METHODE","CUST_INDUSTRY","BIG_GROUPING_CUST","PIC_NAME_NEW RELATION","PIC SUPPORT DATA","UNIT",
            "DEPT","DATE","PERIODE","PERIODE_WEEK","ORIGIN RT","DEST RT","3LC Last Status","Regional Last Status",
            "Zona Last Status","OTS by CT","PERIODE_OTS","CLOSING_OTS","CODING_RT","RECEIVED/REASON_RT","Return Date"
        ]

    def parse_yy_mm(self, inp):
        tahun = int("20" + inp[:2])
        bulan = int(inp[2:])
        return tahun, bulan

    def month_iter(self, start, end):
        cur = start
        while cur <= end:
            yield cur
            if cur.month == 12:
                cur = datetime(cur.year + 1, 1, 1)
            else:
                cur = datetime(cur.year, cur.month + 1, 1)

    def get_criteria_for_project(self, proj_name):
        if not self.criteria_lists:
            return None
        for crit in self.criteria_lists:
            group_name = crit.get("group_name") or crit.get("groups_name")
            if not group_name:
                continue
            if group_name.strip().lower() == proj_name.strip().lower():
                return crit
        return None

    def run(self, request_id=None, filter_categories=None):
        start_year, start_month = self.parse_yy_mm(self.start_yy_mm)
        end_year, end_month = self.parse_yy_mm(self.end_yy_mm)

        if self.start_date and self.end_date:
            filter_start = self.start_date
            filter_end_exclusive = self.end_date + timedelta(days=1)
            start_month_date = datetime(self.start_date.year, self.start_date.month, 1)
            end_month_date = datetime(self.end_date.year, self.end_date.month, 1)
        else:
            start_month_date = datetime(start_year, start_month, 1)
            end_month_date = datetime(end_year, end_month, 1)
            filter_start = start_month_date
            if end_month_date.month == 12:
                filter_end_exclusive = datetime(end_month_date.year + 1, 1, 1)
            else:
                filter_end_exclusive = datetime(end_month_date.year, end_month_date.month + 1, 1)

        # ======================================================
        # 🔹 Tentukan sumber kategori (berdasarkan mode & filter)
        # ======================================================
        if self.full:
            if filter_categories:
                # ✅ FULL MODE + filter aktif → pakai filter user
                categories = [c.lower().strip() for c in filter_categories]
                print(f"📂 [FULL MODE] Gunakan kategori dari filter: {categories}")
            else:
                # ✅ FULL MODE tanpa filter → pakai kategori dari project terpilih saja
                categories = list({proj["category"].lower() for proj in self.project_lists})
                print(f"📘 [FULL MODE] Tanpa filter, gunakan kategori project: {categories}")
        else:
            # PERFORMANCE / REPORT MODE → tetap ambil dari project reference
            categories = list({proj["category"].lower() for proj in self.project_lists})
            print(f"📘 [PERFORMANCE MODE] Kategori dari project reference: {categories}")

        project_id_map = {
            proj["name"]: [str(x).replace("'", "").strip() for x in proj.get("id_account", [])]
            for proj in self.project_lists
        }
        results = {proj["name"]: [] for proj in self.project_lists}

        # ======================================================
        # 🔹 Initialize Progress Tracking
        # ======================================================
        STAGE_WEIGHTS = {"reading": 0.4, "processing": 0.4, "saving": 0.2}

        # ======================================================
        # 🔹 Normalize & apply category filter (Full Mode only)
        # ======================================================
        if filter_categories:
            filter_categories = [c.lower().strip() for c in filter_categories]
            print(f"🔍 [FULL MODE] Filter kategori aktif: {filter_categories}")

        # ======================================================
        # 🔹 Kumpulkan semua file
        # ======================================================
        all_files = []
        for cur_date in self.month_iter(start_month_date, end_month_date):
            year, month = cur_date.year, cur_date.month
            month_name = self.bulan_id[month]
            folder_base = fr"{self.base_path}\{year}\{month}. {month_name} {year}\CATEGORY"
            print(f"▶ Proses: {month_name} {year}")

            for category in categories:
                # ✅ Lewati kategori yang tidak dipilih (jika filter aktif)
                if filter_categories and category.lower() not in filter_categories:
                    print(f"   ⚪ Lewati kategori {category.upper()} (tidak dipilih)")
                    continue

                folder_path = os.path.join(folder_base, category.upper())
                if not os.path.exists(folder_path):
                    print(f"   ⚠ Folder kategori {category.upper()} tidak ditemukan.")
                    continue

                print(f"   ✅ Proses kategori: {category.upper()}")

                for file_name in os.listdir(folder_path):
                    if not file_name.endswith('.csv'):
                        continue

                    fname_lower = file_name.lower()
                    if self.status and self.status.lower() != "all":
                        if self.status.lower() == "close" and "close" not in fname_lower:
                            continue
                        if self.status.lower() == "open" and "open" not in fname_lower:
                            continue

                    all_files.append((folder_path, file_name))

        total_files = len(all_files)
        if total_files == 0:
            print("⚠ Tidak ada file CSV ditemukan.")
            if request_id and self.progress_dict is not None:
                self.progress_dict[request_id] = {"current": 100, "total": 100}
            return []

        # ======================================================
        # 🔹 Hitung total chunks (estimasi)
        # ======================================================
        chunksize = 50000
        total_chunks = 0
        for folder_path, file_name in all_files:
            file_path = os.path.join(folder_path, file_name)
            try:
                df_head = _read_csv(file_path, nrows=5)
                file_size = os.path.getsize(file_path)
                avg_row_size = file_size / (len(df_head) + 1) if len(df_head) > 0 else 1000
                estimated_rows = file_size / avg_row_size
                total_chunks += max(1, (int(estimated_rows) + chunksize - 1) // chunksize)
            except Exception as e:
                print(f"Error estimating chunks for {file_path}: {e}")
                total_chunks += 1

        total_projects = len(results)
        total_saves = total_projects
        total_tasks = (
            total_files * STAGE_WEIGHTS["reading"]
            + total_chunks * STAGE_WEIGHTS["processing"]
            + total_saves * STAGE_WEIGHTS["saving"]
        )

        if request_id and self.progress_dict is not None:
            self.progress_dict[request_id] = {"current": 0, "total": 100}
            current_progress = 0.0

        # ======================================================
        # 🔹 Loop semua file (Reading + Processing)
        # ======================================================
        processed_files = 0
        processed_chunks = 0

        for i, (folder_path, file_name) in enumerate(all_files, start=1):
            file_path = os.path.join(folder_path, file_name)
            print(f"   ⏳ Baca file: {file_path}")

            try:
                chunk_iterator = _read_csv(file_path, chunksize=chunksize, low_memory=False)
            except Exception as e:
                print(f"Error reading {file_path}: {e}")
                continue

            for chunk in chunk_iterator:
                if "TGL_ENTRY" in chunk.columns:
                    entry_dates = pd.to_datetime(chunk["TGL_ENTRY"], errors="coerce")
                    chunk = chunk[(entry_dates >= filter_start) & (entry_dates < filter_end_exclusive)]
                    if chunk.empty:
                        continue

                # ======================================================
                # 🟢 FILTER BY CATEGORY (Full Mode Only - di level data)
                # ======================================================
                if self.full and filter_categories and "CATEGORY" in chunk.columns:
                    chunk = chunk[chunk["CATEGORY"].str.lower().isin(filter_categories)]
                    if chunk.empty:
                        continue

                # ======================================================
                # 🔹 PROSES PER PROJECT (default)
                # ======================================================
                for proj in self.project_lists:
                    proj_name = proj["name"]
                    id_accounts_clean = project_id_map[proj_name]

                    df_split = chunk[
                        chunk['ID_ACCOUNT'].astype(str).str.replace("'", "").isin(id_accounts_clean)
                    ].copy()
                    if df_split.empty:
                        continue

                    if proj.get("late_delivery", False):
                        df_split = normalize_all_dates(df_split, debug=False)
                        df_split = add_grouping_late(df_split)

                    if not self.report or (self.status and self.status.lower() not in ("", "all")):
                        df_split = add_status_pod_2(df_split, debug=False)
                    if 'STATUS_POD_2' not in df_split.columns:
                        print(f"Warning: STATUS_POD_2 missing in chunk from {file_path}")
                        df_split['STATUS_POD_2'] = 'Unknown'

                    # 🔹 Filter berdasarkan status (close/open)
                    if self.status and self.status.lower() != "all":
                        if self.status.lower() == "close":
                            df_split = df_split[df_split["STATUS_POD_2"].isin(["Success", "Return Shipper"])]
                        elif self.status.lower() == "open":
                            df_split = df_split[~df_split["STATUS_POD_2"].isin(["Success", "Return Shipper"])]

                    # 🔹 Pilih kolom minimal (hanya jika bukan full/report)
                    if not self.full and not self.report:
                        cols = [col for col in self.columns_to_extract if col in df_split.columns]
                        df_split = df_split[cols]

                    results[proj_name].append(df_split)

                processed_chunks += 1
                del chunk, df_split

            processed_files += 1
            if request_id and self.progress_dict is not None:
                reading_progress = (processed_files / total_files) * STAGE_WEIGHTS["reading"] * 100
                processing_progress = (processed_chunks / total_chunks) * STAGE_WEIGHTS["processing"] * 100
                current_progress = reading_progress + processing_progress
                self.progress_dict[request_id]["current"] = min(round(current_progress), 99)

        # ======================
        # 🔹 Simpan hasil ke Excel (Saving Stage)
        # ======================
        MAX_ROWS_EXCEL = 1_000_000
        saved_files = []
        processed_saves = 0

        # Perbarui total_saves berdasarkan data aktual
        total_saves = 0
        for proj_name, df_list in results.items():
            if not df_list:
                continue
            df_final = pd.concat(df_list, ignore_index=True)

            # 🔹 Hapus AWB duplikat, prioritaskan TGL_TARIK_REPORT terbaru
            if "AWB" in df_final.columns:
                if "TGL_TARIK_REPORT" in df_final.columns:
                    df_final["TGL_TARIK_REPORT"] = pd.to_datetime(
                        df_final["TGL_TARIK_REPORT"], errors="coerce"
                    )
                    df_final = df_final.sort_values(
                        by=["AWB", "TGL_TARIK_REPORT"], ascending=[True, False]
                    )
                df_final = df_final.drop_duplicates(subset=["AWB"], keep="last")

            total_rows = len(df_final)
            total_saves += (total_rows // MAX_ROWS_EXCEL) + (1 if total_rows % MAX_ROWS_EXCEL else 0)
        if total_saves == 0:
            total_saves = total_projects  # Fallback jika tidak ada data

        for proj_name, df_list in results.items():
            if not df_list:
                print(f"⚠ Tidak ada data untuk project {proj_name}")
                continue

            df_final = pd.concat(df_list, ignore_index=True)

            # 🔹 Hapus AWB duplikat, prioritaskan TGL_TARIK_REPORT terbaru
            if "AWB" in df_final.columns:
                if "TGL_TARIK_REPORT" in df_final.columns:
                    df_final["TGL_TARIK_REPORT"] = pd.to_datetime(
                        df_final["TGL_TARIK_REPORT"], errors="coerce"
                    )
                    df_final = df_final.sort_values(
                        by=["AWB", "TGL_TARIK_REPORT"], ascending=[True, False]
                    )
                df_final = df_final.drop_duplicates(subset=["AWB"], keep="last")

            # Normalisasi semua kolom tanggal
            df_final = normalize_all_dates(df_final, debug=False)

            # --- Jalankan transform hanya jika perlu ---
            if not self.full:
                criteria = self.get_criteria_for_project(proj_name)
                if criteria and "selected_cols" in criteria:
                    selected_cols = criteria["selected_cols"]
                    needed_cols = set(selected_cols)

                    # --- Jalankan TRANSFORM_GROUPS jika ada kolomnya di selected_cols ---
                    for tg_name, group in TRANSFORM_GROUPS.items():
                        if any(col in needed_cols for col in group["cols"]):
                            try:
                                df_final = group["func"](df_final, self.base_path)
                                print(f"✅ Group transform {tg_name} dijalankan.")
                            except Exception as e:
                                print(f"⚠️ Gagal transform group {tg_name}: {e}")

                    # --- Jalankan TRANSFORM_FUNCS jika kolom target ada di selected_cols ---
                    for col in needed_cols:
                        if col in TRANSFORM_FUNCS:
                            try:
                                df_final = TRANSFORM_FUNCS[col](df_final, self.base_path)
                                print(f"✅ Transform func {col} dijalankan.")
                            except Exception as e:
                                print(f"⚠️ Gagal transform func {col}: {e}")

                    # --- Pastikan semua kolom ada (isi NaN kalau belum ada) ---
                    for col in selected_cols:
                        if col not in df_final.columns:
                            df_final[col] = ""
                            print(f"ℹ️ Kolom {col} tidak ada di data, dibuat kosong.")

                    # --- Akhirnya filter sesuai selected_cols ---
                    df_final = df_final[[c for c in selected_cols if c in df_final.columns]]
                else:
                    print(f"⚠️ Criteria tidak ditemukan untuk {proj_name}, skip transform.")

            # Tambah kolom aging & career jika belum ada
            if filter_categories:
                print(f"ℹ️ Mode kategori-only — lewati aging_carrer & rename kolom.")
            elif self.criteria_lists:
                print(f"ℹ️ Mode report aktif — lewati aging_carrer.")
            else:
                df_final = add_aging_carrer(df_final)
                df_final.rename(columns=self.renamed_columns, inplace=True)
            
            # Perbaiki kolom regional jika perlu
            df_final = fix_regional_cols(df_final)

            subfolder = os.path.join(self.archive_dir, proj_name)
            os.makedirs(subfolder, exist_ok=True)

            total_rows = len(df_final)
            num_parts = (total_rows // MAX_ROWS_EXCEL) + (1 if total_rows % MAX_ROWS_EXCEL else 0)

            safe_proj_name = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in proj_name)

            for part in range(num_parts):
                start = part * MAX_ROWS_EXCEL
                end = min((part + 1) * MAX_ROWS_EXCEL, total_rows)
                df_part = df_final.iloc[start:end]

                suffix = f"({part})" if part > 0 else ""
                filename = f"{safe_proj_name}{suffix} {self.start_yy_mm}-{self.end_yy_mm}"
                
                if self.full:
                    # 🔹 Simpan sebagai CSV
                    save_path = os.path.join(subfolder, filename + ".csv")
                    try:
                        # Pastikan semua kolom di full_columns ada
                        for col in self.full_columns:
                            if col not in df_part.columns:
                                df_part[col] = pd.NA

                        # Urutkan kolom sesuai definisi full_columns
                        df_part = df_part[self.full_columns]

                        df_part.to_csv(save_path, index=False, encoding="utf-8-sig")
                        print(f"✅ [FULL MODE] Project {proj_name} part {part+1}/{num_parts} disimpan ke CSV: {save_path}")
                        saved_files.append(save_path)
                    except Exception as e:
                        print(f"Error saving CSV {save_path}: {e}")
                        continue
                else:
                    # 🔹 Simpan sebagai Excel
                    save_path = os.path.join(subfolder, filename + ".xlsx")
                    try:
                        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                            df_part.to_excel(writer, sheet_name=proj_name, index=False)

                        wb = openpyxl.load_workbook(save_path)
                        ws = wb[proj_name]
                        nrows, ncols = ws.max_row, ws.max_column
                        last_col = openpyxl.utils.get_column_letter(ncols)
                        table_range = f"A1:{last_col}{nrows}"

                        # === Mode NORMAL ===
                        if not self.report:
                            table = Table(displayName="DATA_DETAIL", ref=table_range)
                            style = TableStyleInfo(
                                name="TableStyleMedium9",
                                showFirstColumn=False,
                                showLastColumn=False,
                                showRowStripes=True,
                                showColumnStripes=False
                            )
                            table.tableStyleInfo = style
                            ws.add_table(table)

                        # === Mode REPORT ===
                        else:
                            thin_border = Border(
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                top=Side(style="thin"),
                                bottom=Side(style="thin")
                            )
                            header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

                            # Loop semua sel untuk border + format tanggal
                            for row in ws.iter_rows():
                                for cell in row:
                                    # Border
                                    cell.border = thin_border

                                    # Header style
                                    if cell.row == 1:
                                        cell.fill = header_fill

                                    # Format tanggal
                                    if isinstance(cell.value, (datetime, pd.Timestamp)):
                                        cell.number_format = "MM/DD/YYYY"

                            # Auto adjust column width
                            for col_cells in ws.columns:
                                col_letter = get_column_letter(col_cells[0].column)
                                max_length = max(len(str(cell.value) if cell.value else "") for cell in col_cells)
                                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)  # batas agar tidak terlalu lebar

                        # Simpan workbook
                        wb.save(save_path)

                        print(f"✅ Project {proj_name} part {part+1}/{num_parts} disimpan ke Excel: {save_path}")
                        saved_files.append(save_path)

                    except Exception as e:
                        print(f"❌ Gagal menyimpan file {save_path}: {e}")
                        continue

                processed_saves += 1
                if request_id and self.progress_dict is not None:
                    saving_progress = (processed_saves / total_saves) * STAGE_WEIGHTS["saving"] * 100
                    current_progress = (
                        (processed_files / total_files) * STAGE_WEIGHTS["reading"] * 100 +
                        (processed_chunks / total_chunks) * STAGE_WEIGHTS["processing"] * 100 +
                        saving_progress
                    )
                    self.progress_dict[request_id]["current"] = min(round(current_progress), 100)

        # Set progress ke 100% setelah semua selesai
        if request_id and self.progress_dict is not None:
            self.progress_dict[request_id]["current"] = 100

        print("🎉 All data processed and saved successfully.")
        return saved_files
